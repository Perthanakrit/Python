from openpyxl import load_workbook

workbook = load_workbook(filename="score2.xlsx")
workbook.sheetnames

sheet = workbook.active

scorelist = []
passed_exam = []
failed_exam = []


def score_result():
    score_total = 0
    for rowNum in range(5, 25):
        sheet.cell(rowNum, 8).value = 0
        for i in range(6):  # loop for 0-5
            sheet.cell(rowNum, 8).value += sheet.cell(rowNum, i + 2).value
        # finding the number of people who passed and failed the exam
        if sheet.cell(rowNum, 7).value >= 15:
            passed_exam.append(sheet.cell(rowNum, 7).value)
        elif sheet.cell(rowNum, 7).value < 15:
            failed_exam.append(sheet.cell(rowNum, 7).value)
        #
        score = sheet.cell(rowNum, 8).value
        if score >= 80:
            sheet.cell(rowNum, 9).value = 'A'
        elif 70 <= score < 80:
            sheet.cell(rowNum, 9).value = 'B'
        elif 60 <= score < 70:
            sheet.cell(rowNum, 9).value = 'C'
        elif 50 <= score < 60:
            sheet.cell(rowNum, 9).value = 'D'
        elif score < 50:
            sheet.cell(rowNum, 9).value = 'F'

        score_total += score
        print(sheet.cell(rowNum, 8).value, sheet.cell(rowNum, 9).value)
        scorelist.append(score)
    sheet['L13'] = score_total / len(range(5, 25))
    print(sheet['L13'].value)
    sheet['L14'], sheet['L15'] = len(passed_exam), len(failed_exam)
    print('Passed and Failed', sheet['L14'].value, sheet['L15'].value)
    return scorelist


def score(Para_list):
    sheet['L11'] = min(Para_list)
    sheet['L12'] = max(Para_list)
    return sheet['L11'].value, sheet['L12'].value


score_result()
print(score(scorelist)[0], score(scorelist)[1])

workbook.save(filename="score2.xlsx")